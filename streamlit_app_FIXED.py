import streamlit as st
import pandas as pd
import numpy as np
import io
from datetime import datetime
from openpyxl import load_workbook

# Page configuration
st.set_page_config(
    page_title="MLS vs CAMA Data Comparison",
    page_icon="📊",
    layout="wide"
)

# --- Configuration ---
UNIQUE_ID_COLUMN = {'mls_col': 'Parcel Number', 'cama_col': 'PARID'}

COLUMNS_TO_COMPARE = [
    {'mls_col': 'Above Grade Finished Area', 'cama_col': 'SFLA'},
    {'mls_col': 'Bedrooms Total', 'cama_col': 'RMBED'},
    {'mls_col': 'Bathrooms Full', 'cama_col': 'FIXBATH'},
    {'mls_col': 'Bathrooms Half', 'cama_col': 'FIXHALF'},
]

COLUMNS_TO_COMPARE_SUM = [
    {'mls_col': 'Below Grade Finished Area', 'cama_cols': ['RECROMAREA', 'FINBSMTAREA', 'UFEATAREA']}
]

COLUMNS_TO_COMPARE_CATEGORICAL = [
    {
        'mls_col': 'Cooling',
        'cama_col': 'HEAT',
        'mls_check_contains': 'Central Air',
        'cama_expected_if_true': 1,
        'cama_expected_if_false': 0,
        'case_sensitive': False
    }
]

NUMERIC_TOLERANCE = 0.01
SKIP_ZERO_VALUES = True

ADDRESS_COLUMNS = {
    'address': 'Address',
    'city': 'City',
    'state': 'State or Province',
    'zip': 'Postal Code'
}

ZILLOW_URL_BASE = "https://www.zillow.com/homes/"

# --- Helper Functions ---

def format_zillow_url(address, city, state, zip_code):
    """Create a Zillow search URL from address components."""
    if pd.isna(address) or pd.isna(city) or pd.isna(zip_code):
        return None
    
    import re
    
    address_clean = str(address).strip()
    city_clean = str(city).strip()
    zip_clean = str(zip_code).strip().split('-')[0]
    
    address_clean = re.sub(r'\s+(Apt|Unit|#|Suite)\s*[\w-]*$', '', address_clean, flags=re.IGNORECASE)
    address_formatted = re.sub(r'[^\w\s-]', '', address_clean)
    address_formatted = re.sub(r'\s+', '-', address_formatted)
    
    city_formatted = re.sub(r'[^\w\s-]', '', city_clean)
    city_formatted = re.sub(r'\s+', '-', city_formatted)
    
    url_slug = f"{address_formatted}-{city_formatted}-OH-{zip_clean}_rb"
    return f"{ZILLOW_URL_BASE}{url_slug}/"

def values_equal(val1, val2):
    """Check if two values are equal within tolerance."""
    try:
        num1 = pd.to_numeric(val1, errors='raise')
        num2 = pd.to_numeric(val2, errors='raise')

        if pd.isna(num1) and pd.isna(num2):
            return True
        elif pd.isna(num1) != pd.isna(num2):
            return False
        else:
            return np.isclose(num1, num2, equal_nan=False, rtol=1e-9, atol=NUMERIC_TOLERANCE)
    except (ValueError, TypeError):
        str1 = str(val1).strip().lower() if pd.notna(val1) else ''
        str2 = str(val2).strip().lower() if pd.notna(val2) else ''
        return str1 == str2

def categorical_match(mls_val, cama_val, mapping):
    """Check if categorical MLS field matches expected CAMA value."""
    check_text = mapping.get('mls_check_contains', '')
    expected_if_true = mapping.get('cama_expected_if_true')
    expected_if_false = mapping.get('cama_expected_if_false')
    case_sensitive = mapping.get('case_sensitive', False)

    mls_str = str(mls_val).strip() if pd.notna(mls_val) else ''

    if not case_sensitive:
        mls_str = mls_str.lower()
        check_text = check_text.lower()

    text_found = check_text in mls_str
    expected_cama = expected_if_true if text_found else expected_if_false

    try:
        cama_numeric = pd.to_numeric(cama_val, errors='coerce')
        expected_numeric = pd.to_numeric(expected_cama, errors='coerce')

        if pd.isna(cama_numeric) and pd.isna(expected_numeric):
            return True
        elif pd.isna(cama_numeric) or pd.isna(expected_numeric):
            return False
        else:
            return np.isclose(cama_numeric, expected_numeric, equal_nan=False, rtol=1e-9, atol=NUMERIC_TOLERANCE)
    except:
        return str(cama_val).strip().lower() == str(expected_cama).strip().lower()

def calculate_difference(val1, val2):
    """Calculate the difference between two values."""
    try:
        num1 = pd.to_numeric(val1, errors='raise')
        num2 = pd.to_numeric(val2, errors='raise')

        if pd.isna(num1) or pd.isna(num2):
            return "N/A"

        diff = num1 - num2
        return f"{diff:,.2f}"
    except (ValueError, TypeError):
        return "Text difference"

def compare_data_enhanced(df_mls, df_cama, unique_id_col, cols_to_compare_mapping,
                         cols_to_compare_sum=None, cols_to_compare_categorical=None, 
                         window_id=None):
    """Compare MLS and CAMA dataframes and return discrepancies."""
    
    if df_mls is None or df_cama is None:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    mls_id_col_name = unique_id_col.get('mls_col')
    cama_id_col_name = unique_id_col.get('cama_col')

    if mls_id_col_name not in df_mls.columns:
        st.error(f"Column '{mls_id_col_name}' not found in MLS data")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    if cama_id_col_name not in df_cama.columns:
        st.error(f"Column '{cama_id_col_name}' not found in CAMA data")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    # Rename and merge
    df_mls_renamed = df_mls.copy()
    df_mls_renamed = df_mls_renamed.rename(columns={mls_id_col_name: cama_id_col_name})
    
    matched_df = pd.merge(df_mls_renamed, df_cama, on=cama_id_col_name, how='inner')
    merged_df = pd.merge(df_mls_renamed, df_cama, on=cama_id_col_name, how='outer', indicator=True)

    missing_in_cama = []
    missing_in_mls = []
    value_mismatches = []
    perfect_matches = []

    # Build parcel URL template if window_id provided
    if window_id:
        parcel_url_template = f"https://iasworld.starkcountyohio.gov/iasworld/Maintain/Transact.aspx?txtMaskedPin={{parcel_id}}&selYear=&userYear=&selJur=&chkShowHistory=False&chkShowChanges=&chkShowDeactivated=&PinValue={{parcel_id}}&pin=&trans_key=&windowId={window_id}&submitFlag=true&TransPopUp=&ACflag=False&ACflag2=False"
    else:
        parcel_url_template = None

    for index, row in merged_df.iterrows():
        record_id = row.get(cama_id_col_name)
        merge_status = row.get('_merge')

        if merge_status == 'left_only':
            listing_num = row.get('Listing #', '')
            closed_date = row.get('Closed Date', '')
            missing_in_cama.append({
                'Parcel_ID': record_id,
                'Listing_Number': listing_num,
                'Closed_Date': closed_date
            })

        elif merge_status == 'right_only':
            missing_in_mls.append({'Parcel_ID': record_id})

        elif merge_status == 'both':
            listing_num = row.get('Listing #', '')
            salekey = row.get('SALEKEY', '')
            nopar = row.get('NOPAR', '')
            additional_parcels = row.get('ADDITIONAL_PARCELS', '')
            
            address = row.get(ADDRESS_COLUMNS.get('address', 'Address'), '')
            city = row.get(ADDRESS_COLUMNS.get('city', 'City'), '')
            state = row.get(ADDRESS_COLUMNS.get('state', 'State or Province'), '')
            zip_code = row.get(ADDRESS_COLUMNS.get('zip', 'Postal Code'), '')
            
            record_mismatches = []
            fields_compared = []

            # Standard comparisons
            for mapping in cols_to_compare_mapping:
                mls_col = mapping['mls_col']
                cama_col = mapping['cama_col']

                if mls_col not in merged_df.columns or cama_col not in merged_df.columns:
                    continue

                mls_val = row.get(mls_col)
                cama_val = row.get(cama_col)

                mls_is_blank = pd.isna(mls_val) or (isinstance(mls_val, str) and mls_val.strip() == '')
                cama_is_blank = pd.isna(cama_val) or (isinstance(cama_val, str) and cama_val.strip() == '')

                if mls_is_blank or cama_is_blank:
                    continue

                fields_compared.append(mls_col)

                if SKIP_ZERO_VALUES:
                    try:
                        mls_numeric = pd.to_numeric(mls_val, errors='coerce')
                        cama_numeric = pd.to_numeric(cama_val, errors='coerce')
                        # Only skip if BOTH values are zero (both agree there's nothing)
                        # If one is zero and the other is not, that's a mismatch!
                        if (pd.notna(mls_numeric) and mls_numeric == 0) and (pd.notna(cama_numeric) and cama_numeric == 0):
                            continue
                    except:
                        pass

                if not values_equal(mls_val, cama_val):
                    record_mismatches.append({
                        'Parcel_ID': record_id,
                        'NOPAR': nopar,
                        'ADDITIONAL_PARCELS': additional_parcels,
                        'Listing_Number': listing_num,
                        'SALEKEY': salekey,
                        'Address': address,
                        'City': city,
                        'State': state,
                        'Zip': zip_code,
                        'Field_MLS': mls_col,
                        'Field_CAMA': cama_col,
                        'MLS_Value': mls_val,
                        'CAMA_Value': cama_val,
                        'Difference': calculate_difference(mls_val, cama_val),
                        'Parcel_URL': parcel_url_template.format(parcel_id=record_id) if parcel_url_template else '',
                        'Zillow_URL': format_zillow_url(address, city, state, zip_code)
                    })

            # Sum comparisons
            if cols_to_compare_sum:
                for mapping in cols_to_compare_sum:
                    mls_col = mapping['mls_col']
                    cama_cols = mapping['cama_cols']

                    if mls_col not in merged_df.columns:
                        continue
                    
                    missing_cols = [col for col in cama_cols if col not in merged_df.columns]
                    if missing_cols:
                        continue

                    mls_val = row.get(mls_col)
                    mls_is_blank = pd.isna(mls_val) or (isinstance(mls_val, str) and mls_val.strip() == '')
                    if mls_is_blank:
                        continue

                    cama_sum = 0
                    all_cama_blank = True
                    for col in cama_cols:
                        val = row.get(col)
                        if pd.notna(val):
                            all_cama_blank = False
                            try:
                                cama_sum += pd.to_numeric(val, errors='coerce')
                            except:
                                pass

                    if all_cama_blank:
                        continue

                    fields_compared.append(mls_col)

                    if SKIP_ZERO_VALUES:
                        try:
                            mls_numeric = pd.to_numeric(mls_val, errors='coerce')
                            # Only skip if BOTH values are zero (both agree there's no below-grade area)
                            # If one is zero and the other is not, that's a mismatch!
                            if (pd.notna(mls_numeric) and mls_numeric == 0) and (cama_sum == 0):
                                continue
                        except:
                            pass

                    if not values_equal(mls_val, cama_sum):
                        record_mismatches.append({
                            'Parcel_ID': record_id,
                            'NOPAR': nopar,
                            'ADDITIONAL_PARCELS': additional_parcels,
                            'Listing_Number': listing_num,
                            'SALEKEY': salekey,
                            'Address': address,
                            'City': city,
                            'State': state,
                            'Zip': zip_code,
                            'Field_MLS': mls_col,
                            'Field_CAMA': f"SUM({', '.join(cama_cols)})",
                            'MLS_Value': mls_val,
                            'CAMA_Value': cama_sum,
                            'Difference': calculate_difference(mls_val, cama_sum),
                            'Parcel_URL': parcel_url_template.format(parcel_id=record_id) if parcel_url_template else '',
                            'Zillow_URL': format_zillow_url(address, city, state, zip_code)
                        })

            # Categorical comparisons
            if cols_to_compare_categorical:
                for mapping in cols_to_compare_categorical:
                    mls_col = mapping['mls_col']
                    cama_col = mapping['cama_col']

                    if mls_col not in merged_df.columns or cama_col not in merged_df.columns:
                        continue

                    mls_val = row.get(mls_col)
                    cama_val = row.get(cama_col)

                    mls_is_blank = pd.isna(mls_val) or (isinstance(mls_val, str) and mls_val.strip() == '')
                    cama_is_blank = pd.isna(cama_val) or (isinstance(cama_val, str) and cama_val.strip() == '')

                    if mls_is_blank or cama_is_blank:
                        continue

                    fields_compared.append(mls_col)
                    is_match = categorical_match(mls_val, cama_val, mapping)

                    check_text = mapping.get('mls_check_contains', '')
                    case_sensitive = mapping.get('case_sensitive', False)
                    mls_str = str(mls_val).strip() if pd.notna(mls_val) else ''
                    if not case_sensitive:
                        mls_str = mls_str.lower()
                        check_text_lower = check_text.lower()
                    else:
                        check_text_lower = check_text
                    text_found = check_text_lower in mls_str
                    expected_cama = mapping.get('cama_expected_if_true') if text_found else mapping.get('cama_expected_if_false')

                    if not is_match:
                        record_mismatches.append({
                            'Parcel_ID': record_id,
                            'NOPAR': nopar,
                            'ADDITIONAL_PARCELS': additional_parcels,
                            'Listing_Number': listing_num,
                            'SALEKEY': salekey,
                            'Address': address,
                            'City': city,
                            'State': state,
                            'Zip': zip_code,
                            'Field_MLS': mls_col,
                            'Field_CAMA': cama_col,
                            'MLS_Value': mls_val,
                            'CAMA_Value': cama_val,
                            'Expected_CAMA_Value': expected_cama,
                            'Match_Rule': f"If '{check_text}' in {mls_col}, then {cama_col} should be {mapping.get('cama_expected_if_true')}, else {mapping.get('cama_expected_if_false')}",
                            'Parcel_URL': parcel_url_template.format(parcel_id=record_id) if parcel_url_template else '',
                            'Zillow_URL': format_zillow_url(address, city, state, zip_code)
                        })

            # Perfect matches
            if not record_mismatches and fields_compared:
                perfect_matches.append({
                    'Parcel_ID': record_id,
                    'NOPAR': nopar,
                    'ADDITIONAL_PARCELS': additional_parcels,
                    'Listing_Number': listing_num,
                    'SALEKEY': salekey,
                    'Address': address,
                    'City': city,
                    'State': state,
                    'Zip': zip_code,
                    'Fields_Compared': len(fields_compared),
                    'Fields_List': ', '.join(fields_compared),
                    'Parcel_URL': parcel_url_template.format(parcel_id=record_id) if parcel_url_template else '',
                    'Zillow_URL': format_zillow_url(address, city, state, zip_code)
                })

            value_mismatches.extend(record_mismatches)

    return (pd.DataFrame(missing_in_cama), pd.DataFrame(missing_in_mls), 
            pd.DataFrame(value_mismatches), matched_df, pd.DataFrame(perfect_matches))

def create_excel_with_hyperlinks(df, sheet_name='Sheet1'):
    """Create Excel file with hyperlinks in memory."""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    
    output.seek(0)
    wb = load_workbook(output)
    ws = wb[sheet_name]
    
    # Add Parcel_ID hyperlinks
    if 'Parcel_ID' in df.columns and 'Parcel_URL' in df.columns:
        parcel_col_idx = list(df.columns).index('Parcel_ID') + 1
        url_col_idx = list(df.columns).index('Parcel_URL') + 1
        
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=parcel_col_idx)
            url = ws.cell(row=row_idx, column=url_col_idx).value
            if url and str(url).strip() and str(url) != 'nan':
                cell.hyperlink = url
                cell.style = 'Hyperlink'
    
    # Add Address hyperlinks to Zillow
    if 'Address' in df.columns and 'Zillow_URL' in df.columns:
        address_col_idx = list(df.columns).index('Address') + 1
        zillow_col_idx = list(df.columns).index('Zillow_URL') + 1
        
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=address_col_idx)
            url = ws.cell(row=row_idx, column=zillow_col_idx).value
            if url and str(url).strip() and str(url) != 'nan':
                cell.hyperlink = url
                cell.style = 'Hyperlink'
    
    # Remove URL columns (they were just for creating hyperlinks)
    if 'Parcel_URL' in df.columns:
        url_col_idx = list(df.columns).index('Parcel_URL') + 1
        ws.delete_cols(url_col_idx)
    
    if 'Zillow_URL' in df.columns:
        # Recalculate index after potential deletion
        remaining_cols = [col for col in df.columns if col not in ['Parcel_URL']]
        if 'Zillow_URL' in remaining_cols:
            zillow_col_idx = remaining_cols.index('Zillow_URL') + 1
            ws.delete_cols(zillow_col_idx)
    
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    return final_output.getvalue()

def create_zip_with_all_reports(df_missing_cama, df_missing_mls, df_value_mismatches, 
                                 df_perfect_matches, city_comparison_df=None):
    """Create a ZIP file containing all Excel reports and stats CSV with timestamped filenames."""
    import zipfile
    
    # Get current timestamp for filenames (date only)
    timestamp = datetime.now().strftime("%Y-%m-%d")
    
    # Create ZIP file in memory
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Add each Excel report if it has data
        if not df_missing_cama.empty:
            excel_data = create_excel_with_hyperlinks(df_missing_cama, 'Missing in CAMA')
            filename = f"missing_in_CAMA_{timestamp}.xlsx"
            zip_file.writestr(filename, excel_data)
        
        if not df_missing_mls.empty:
            excel_data = create_excel_with_hyperlinks(df_missing_mls, 'Missing in MLS')
            filename = f"missing_in_MLS_{timestamp}.xlsx"
            zip_file.writestr(filename, excel_data)
        
        if not df_value_mismatches.empty:
            excel_data = create_excel_with_hyperlinks(df_value_mismatches, 'Value Mismatches')
            filename = f"value_mismatches_{timestamp}.xlsx"
            zip_file.writestr(filename, excel_data)
        
        if not df_perfect_matches.empty:
            excel_data = create_excel_with_hyperlinks(df_perfect_matches, 'Perfect Matches')
            filename = f"perfect_matches_{timestamp}.xlsx"
            zip_file.writestr(filename, excel_data)
        
        # Add city statistics CSV if available
        if city_comparison_df is not None and not city_comparison_df.empty:
            csv_data = city_comparison_df.to_csv(index=False)
            filename = f"city_match_statistics_{timestamp}.csv"
            zip_file.writestr(filename, csv_data)
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

# --- Streamlit App ---

st.title("📊 MLS vs CAMA Data Comparison Tool")
st.markdown("Compare MLS and CAMA property data to identify discrepancies and perfect matches.")

# Sidebar configuration
with st.sidebar:
    st.header("⚙️ Configuration")
    
    st.subheader("WindowId Setup")
    st.markdown("""
    **How to get WindowId:**
    1. Go to [Stark County iasWorld](https://iasworld.starkcountyohio.gov/iasworld/)
    2. Log in and search for any property
    3. Copy the `windowId` value from the URL
    
    Example: `...windowId=638981240146803746&...`
    """)
    
    window_id = st.text_input(
        "Enter WindowId",
        value="638981240146803746",
        help="Used to generate clickable links to property records"
    )
    
    st.divider()
    
    st.subheader("Comparison Settings")
    tolerance = st.number_input(
        "Numeric Tolerance",
        value=0.01,
        format="%.4f",
        help="Absolute tolerance for numeric comparisons"
    )
    
    skip_zeros = st.checkbox(
        "Skip Zero Values",
        value=True,
        help="If enabled, treats 0 as 'no data' and skips comparison"
    )

# Update global settings
NUMERIC_TOLERANCE = tolerance
SKIP_ZERO_VALUES = skip_zeros

# File upload section
st.header("📁 Upload Data Files")

col1, col2 = st.columns(2)

with col1:
    mls_file = st.file_uploader(
        "Upload MLS Data (Excel)",
        type=['xlsx', 'xls'],
        help="Upload your MLS Excel file"
    )

with col2:
    cama_file = st.file_uploader(
        "Upload CAMA Data (Excel)",
        type=['xlsx', 'xls'],
        help="Upload your CAMA Excel file"
    )

# Process data when both files are uploaded
if mls_file and cama_file:
    
    with st.spinner("Loading data files..."):
        try:
            df_mls = pd.read_excel(mls_file)
            df_cama = pd.read_excel(cama_file)
            st.success("✅ Data files loaded successfully!")
        except Exception as e:
            st.error(f"Error loading files: {e}")
            st.stop()
    
    # Display data summary
    st.header("📊 Data Summary")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("MLS Records", len(df_mls))
    with col2:
        st.metric("CAMA Records", len(df_cama))
    with col3:
        st.metric("Numeric Tolerance", f"{NUMERIC_TOLERANCE}")
    
    # Run comparison button
    if st.button("🔍 Run Comparison", type="primary", use_container_width=True):
        
        with st.spinner("Comparing data... This may take a moment."):
            df_missing_cama, df_missing_mls, df_value_mismatches, matched_df, df_perfect_matches = \
                compare_data_enhanced(
                    df_mls, df_cama, 
                    UNIQUE_ID_COLUMN,
                    COLUMNS_TO_COMPARE,
                    cols_to_compare_sum=COLUMNS_TO_COMPARE_SUM,
                    cols_to_compare_categorical=COLUMNS_TO_COMPARE_CATEGORICAL,
                    window_id=window_id
                )
        
        # Display results
        st.header("📈 Results Summary")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("✅ Matched Records", len(matched_df))
        with col2:
            st.metric("❌ Missing in CAMA", len(df_missing_cama))
        with col3:
            st.metric("❌ Missing in MLS", len(df_missing_mls))
        with col4:
            st.metric("⚠️ Value Mismatches", len(df_value_mismatches))
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("✅ Perfect Matches", len(df_perfect_matches))
        with col2:
            if not df_value_mismatches.empty:
                unique_fields = df_value_mismatches['Field_MLS'].nunique()
                st.metric("📊 Fields with Mismatches", unique_fields)
        
        # Match rate statistics
        st.header("📊 CAMA Parcel Match Statistics")
        st.markdown("Analysis of which CAMA parcels were found in the MLS data")
        
        # Calculate overall match rate
        total_cama_parcels = len(df_cama)
        matched_parcels = len(matched_df)
        match_rate = (matched_parcels / total_cama_parcels * 100) if total_cama_parcels > 0 else 0
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total CAMA Parcels", f"{total_cama_parcels:,}")
        with col2:
            st.metric("Found in MLS", f"{matched_parcels:,}")
        with col3:
            st.metric("Match Rate", f"{match_rate:.2f}%")
        
        # Match rate by city
        st.subheader("Match Rate by City")
        
        # Determine which city column to use - prefer CAMA's city column
        cama_city_col = None
        if 'CITYNAME' in df_cama.columns:
            cama_city_col = 'CITYNAME'
        elif 'City' in df_cama.columns:
            cama_city_col = 'City'
        
        if cama_city_col and not matched_df.empty:
            # Get CAMA ID column
            cama_id_col = UNIQUE_ID_COLUMN.get('cama_col')
            
            # Check if the city column exists in matched_df (it should since it came from CAMA)
            if cama_city_col in matched_df.columns:
                # Get total CAMA parcels by city
                cama_cities = df_cama.groupby(cama_city_col)[cama_id_col].count().reset_index()
                cama_cities.columns = ['City', 'Total_CAMA_Parcels']
                
                # Get matched parcels by city (use the same city column)
                matched_cities = matched_df.groupby(cama_city_col)[cama_id_col].count().reset_index()
                matched_cities.columns = ['City', 'Matched_Parcels']
                
                # Merge the two dataframes
                city_comparison = pd.merge(cama_cities, matched_cities, on='City', how='left')
                city_comparison['Matched_Parcels'] = city_comparison['Matched_Parcels'].fillna(0).astype(int)
                city_comparison['Match_Rate'] = (city_comparison['Matched_Parcels'] / city_comparison['Total_CAMA_Parcels'] * 100).round(2)
                city_comparison['Not_Matched'] = city_comparison['Total_CAMA_Parcels'] - city_comparison['Matched_Parcels']
                
                # Sort by total parcels descending
                city_comparison = city_comparison.sort_values('Total_CAMA_Parcels', ascending=False)
                
                # Store in session state for download all button
                st.session_state['city_comparison'] = city_comparison
                
                # Display as a formatted table
                st.dataframe(
                    city_comparison[['City', 'Total_CAMA_Parcels', 'Matched_Parcels', 'Not_Matched', 'Match_Rate']],
                    use_container_width=True,
                    hide_index=True
                )
                
                # Download button for city statistics
                csv = city_comparison.to_csv(index=False)
                timestamp_csv = datetime.now().strftime("%Y-%m-%d")
                st.download_button(
                    label="📥 Download City Statistics (CSV)",
                    data=csv,
                    file_name=f"city_match_statistics_{timestamp_csv}.csv",
                    mime="text/csv"
                )
                
                # Visualizations
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**Top 10 Cities by Total CAMA Parcels**")
                    top_cities = city_comparison.head(10)
                    st.bar_chart(top_cities.set_index('City')[['Matched_Parcels', 'Not_Matched']])
                
                with col2:
                    st.markdown("**Match Rate by City (Top 10)**")
                    match_rate_chart = top_cities[['City', 'Match_Rate']].set_index('City')
                    st.bar_chart(match_rate_chart)
            else:
                # City column from CAMA not in matched data - try using MLS city
                if 'City' in matched_df.columns:
                    matched_cities = matched_df.groupby('City')[cama_id_col].count().reset_index()
                    matched_cities.columns = ['City', 'Matched_Parcels']
                    matched_cities = matched_cities.sort_values('Matched_Parcels', ascending=False)
                    
                    st.info("ℹ️ City breakdown based on MLS city data (CAMA city field not available in matched records)")
                    st.dataframe(matched_cities, use_container_width=True, hide_index=True)
                else:
                    st.warning("⚠️ City information not available in matched records")
        else:
            st.info("ℹ️ City information not available in the data")
        
        # Mismatch breakdown
        if not df_value_mismatches.empty:
            st.subheader("📊 Mismatches by Field")
            mismatch_counts = df_value_mismatches['Field_MLS'].value_counts()
            st.bar_chart(mismatch_counts)
        
        # Display data previews in tabs
        st.header("📋 Data Preview")
        
        tab1, tab2, tab3, tab4 = st.tabs([
            "Missing in CAMA", 
            "Missing in MLS", 
            "Value Mismatches", 
            "Perfect Matches"
        ])
        
        with tab1:
            if not df_missing_cama.empty:
                st.dataframe(df_missing_cama, use_container_width=True)
            else:
                st.info("No records missing in CAMA")
        
        with tab2:
            if not df_missing_mls.empty:
                st.dataframe(df_missing_mls, use_container_width=True)
            else:
                st.info("No records missing in MLS")
        
        with tab3:
            if not df_value_mismatches.empty:
                st.dataframe(df_value_mismatches, use_container_width=True)
            else:
                st.info("No value mismatches found")
        
        with tab4:
            if not df_perfect_matches.empty:
                st.dataframe(df_perfect_matches, use_container_width=True)
            else:
                st.info("No perfect matches found")
        
        # Download section
        st.header("📥 Download Reports")
        
        # Download All button
        st.markdown("### 📦 Download All Reports")
        
        # Get city comparison if available
        city_comp = st.session_state.get('city_comparison', None)
        
        # Generate timestamp for the ZIP filename (date only)
        timestamp = datetime.now().strftime("%Y-%m-%d")
        zip_filename = f"MLS_CAMA_Comparison_All_Reports_{timestamp}.zip"
        
        # Create the ZIP file
        zip_data = create_zip_with_all_reports(
            df_missing_cama, 
            df_missing_mls, 
            df_value_mismatches, 
            df_perfect_matches,
            city_comp
        )
        
        st.download_button(
            label="📦 Download All Reports (ZIP)",
            data=zip_data,
            file_name=zip_filename,
            mime="application/zip",
            help="Downloads all Excel reports and city statistics in a single ZIP file",
            use_container_width=True
        )
        
        st.markdown("### 📄 Download Individual Reports")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if not df_missing_cama.empty:
                excel_data = create_excel_with_hyperlinks(df_missing_cama, 'Missing in CAMA')
                timestamp = datetime.now().strftime("%Y-%m-%d")
                st.download_button(
                    label="📄 Download Missing in CAMA",
                    data=excel_data,
                    file_name=f"missing_in_CAMA_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            if not df_value_mismatches.empty:
                excel_data = create_excel_with_hyperlinks(df_value_mismatches, 'Value Mismatches')
                timestamp = datetime.now().strftime("%Y-%m-%d")
                st.download_button(
                    label="⚠️ Download Value Mismatches",
                    data=excel_data,
                    file_name=f"value_mismatches_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with col2:
            if not df_missing_mls.empty:
                excel_data = create_excel_with_hyperlinks(df_missing_mls, 'Missing in MLS')
                timestamp = datetime.now().strftime("%Y-%m-%d")
                st.download_button(
                    label="📄 Download Missing in MLS",
                    data=excel_data,
                    file_name=f"missing_in_MLS_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            if not df_perfect_matches.empty:
                excel_data = create_excel_with_hyperlinks(df_perfect_matches, 'Perfect Matches')
                timestamp = datetime.now().strftime("%Y-%m-%d")
                st.download_button(
                    label="✅ Download Perfect Matches",
                    data=excel_data,
                    file_name=f"perfect_matches_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

else:
    st.info("👆 Please upload both MLS and CAMA data files to begin.")
    
    # Show example data format
    with st.expander("ℹ️ Expected Data Format"):
        st.markdown("""
        ### MLS Data Expected Columns:
        - `Parcel Number` (unique identifier)
        - `Above Grade Finished Area`
        - `Bedrooms Total`
        - `Bathrooms Full`
        - `Bathrooms Half`
        - `Below Grade Finished Area`
        - `Cooling`
        - `Address`, `City`, `State or Province`, `Postal Code`
        
        ### CAMA Data Expected Columns:
        - `PARID` (unique identifier)
        - `NOPAR`
        - `CITYNAME` (or `City`) - for city-level statistics
        - `SFLA`
        - `RMBED`
        - `FIXBATH`
        - `FIXHALF`
        - `RECROMAREA`, `FINBSMTAREA`, `UFEATAREA`
        - `HEAT`
        - `SALEKEY`
        
        **Note**: The app automatically detects whether your CAMA data uses 
        `CITYNAME` or `City` for the city column.
        """)

# Footer
st.divider()
st.caption("MLS vs CAMA Comparison Tool | Built with Streamlit")
